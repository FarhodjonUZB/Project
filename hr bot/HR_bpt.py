import telebot
import sqlite3
from telebot import types
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import time
import threading
import os
import uuid
from datetime import datetime

# ================= XAVFSIZLIK VA SOZLAMALAR (PRO) =================
# Token xavfsizligi: Kod ichida ochiq saqlash tavsiya etilmaydi
TOKEN = os.getenv("BOT_TOKEN", "7969875395:AAE8Aw5SpJrfu0kYeKbByNWklk4uixzaD8I")
MAIN_ADMIN_ID = 8356409576
DB_NAME = "hr_bot_v2.db"

# PRO: Savol vaqti limiti (soniya)
QUESTION_TIME = 25 
# Flood nazorati
FLOOD_INTERVAL = 0.8
user_last_req = {}
# Har bir foydalanuvchi uchun savol vaqtini kuzatish
user_question_start_time = {}

# SQLITE LOCK (PRO STABILITY)
db_lock = threading.Lock()

bot = telebot.TeleBot(TOKEN)

# Bazaga ulanish va Optimizatsiya (100k+ userlar uchun)
conn = sqlite3.connect(DB_NAME, check_same_thread=False)
cursor = conn.cursor()

# WAL (Write-Ahead Logging) rejimini yoqish - Bu paralell yozish va o'qishni tezlashtiradi
cursor.execute("PRAGMA journal_mode=WAL;")
# Sinxronizatsiyani optimallashtirish (Xavfsizlik va tezlik balansi)
cursor.execute("PRAGMA synchronous=NORMAL;")
# Kesh hajmini oshirish
cursor.execute("PRAGMA cache_size=10000;")

# ================= DATABASE TABLES =================
with db_lock:
    # 1. Foydalanuvchilar
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS users (
        user_id INTEGER PRIMARY KEY,
        name TEXT,
        phone TEXT,
        registered_at TEXT
    )
    """)
    # Index qo'shish (Qidiruvni tezlashtiradi)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_users_phone ON users(phone)")

    # 2. Savollar
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS questions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        question TEXT,
        media_type TEXT DEFAULT 'text',
        file_id TEXT,
        a TEXT, b TEXT, c TEXT, d TEXT,
        correct TEXT
    )
    """)

    # 3. Javoblar
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS answers (
        user_id INTEGER,
        question_id INTEGER,
        answer TEXT,
        is_correct INTEGER,
        timestamp INTEGER
    )
    """)
    # Index qo'shish: Har safar javob qidirganda butun bazani kovlamaslik uchun
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_answers_user_q ON answers(user_id, question_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_answers_user_correct ON answers(user_id, is_correct)")

    # 4. Sozlamalar
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS settings (
        id INTEGER PRIMARY KEY,
        test_time INTEGER,
        pass_score INTEGER
    )
    """)

    # 5. SESSIONS (PRO DEVICE PROTECTION)
    # Eski test_status o'rniga mukammal sessiya nazorati
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS sessions (
        user_id INTEGER PRIMARY KEY,
        session_id TEXT,
        started_at INTEGER,
        active INTEGER DEFAULT 1,
        warning_sent INTEGER DEFAULT 0
    )
    """)
    # Index: Faol sessiyalarni tez topish uchun
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_sessions_active ON sessions(active)")

    # 6. Adminlar
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS admins (
        user_id INTEGER PRIMARY KEY,
        added_by INTEGER
    )
    """)

    # 7. AUDIT LOGS (PRO LOGGING)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS audit_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        action TEXT,
        timestamp TEXT
    )
    """)
    # Loglar ko'payib ketsa, qidiruv sekinlashmasligi uchun index
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_audit_user ON audit_logs(user_id)")
    
    # 8. Admin logs (Eski loglar uchun)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS admin_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        admin_id INTEGER,
        action TEXT,
        timestamp TEXT
    )
    """)

    # Boshlang'ich sozlamalar
    cursor.execute("INSERT OR IGNORE INTO settings (id, test_time, pass_score) VALUES (1, 600, 70)")
    cursor.execute("INSERT OR IGNORE INTO admins (user_id, added_by) VALUES (?, 0)", (MAIN_ADMIN_ID,))
    conn.commit()

# ================= YORDAMCHI FUNKSIYALAR =================
def is_admin(user_id):
    if user_id == MAIN_ADMIN_ID:
        return True
    with db_lock:
        cursor.execute("SELECT user_id FROM admins WHERE user_id=?", (user_id,))
        return cursor.fetchone() is not None

def log_audit(user_id, action):
    """Barcha muhim harakatlarni loglash"""
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with db_lock:
            cursor.execute("INSERT INTO audit_logs (user_id, action, timestamp) VALUES (?, ?, ?)", 
                        (user_id, action, now))
            conn.commit()
    except Exception as e:
        print(f"Audit log error: {e}")

def log_admin_action(admin_id, action):
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with db_lock:
            cursor.execute("INSERT INTO admin_logs (admin_id, action, timestamp) VALUES (?, ?, ?)", 
                        (admin_id, action, now))
            conn.commit()
    except Exception as e:
        print(f"Admin log error: {e}")

def check_flood(user_id):
    """Spam va Flood himoyasi"""
    now = time.time()
    last = user_last_req.get(user_id, 0)
    user_last_req[user_id] = now
    if now - last < FLOOD_INTERVAL:
        return True
    return False

def get_settings():
    with db_lock:
        cursor.execute("SELECT test_time, pass_score FROM settings WHERE id=1")
        return cursor.fetchone()

# --- SESSION MANAGEMENT (PRO) ---
def create_session(user_id):
    session_id = str(uuid.uuid4())
    with db_lock:
        cursor.execute("""
            INSERT OR REPLACE INTO sessions 
            (user_id, session_id, started_at, active, warning_sent)
            VALUES (?, ?, ?, 1, 0)
        """, (user_id, session_id, int(time.time())))
        conn.commit()
    log_audit(user_id, f"Session Started: {session_id}")
    return session_id

def check_active_session(user_id):
    with db_lock:
        # Index orqali tezkor qidiruv
        cursor.execute("SELECT active FROM sessions WHERE user_id=? AND active=1", (user_id,))
        row = cursor.fetchone()
        return row is not None

def finish_test(user_id, forced=False, reason="Normal"):
    """Testni yakunlash va natijani hisoblash"""
    with db_lock:
        cursor.execute("UPDATE sessions SET active=0 WHERE user_id=?", (user_id,))
        conn.commit()
        
        # Natijani hisoblash (count index orqali tez ishlaydi)
        cursor.execute("SELECT COUNT(*) FROM questions")
        total_questions = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM answers WHERE user_id=? AND is_correct=1", (user_id,))
        correct_answers = cursor.fetchone()[0]
    
    settings = get_settings()
    pass_score = settings[1]
    
    score_percent = 0
    if total_questions > 0:
        score_percent = int((correct_answers / total_questions) * 100)
    
    status_text = "✅ O'TDINGIZ" if score_percent >= pass_score else "❌ YIQILDINGIZ"
    
    msg_text = ""
    if forced:
        msg_text = f"🛑 TEST TO'XTATILDI!\nSabab: {reason}\n\n"
        log_audit(user_id, f"Test Forced Finish: {reason}")
    else:
        msg_text = f"🏁 Test yakunlandi.\n"
        log_audit(user_id, f"Test Finished: {score_percent}%")
        
    msg_text += f"\n📊 Natija: {score_percent}%\n"
    msg_text += f"✅ To'g'ri: {correct_answers}\n"
    msg_text += f"❓ Jami savollar: {total_questions}\n"
    msg_text += f"Hukm: {status_text}"
    
    try:
        bot.send_message(user_id, msg_text, reply_markup=types.ReplyKeyboardRemove(), protect_content=True)
        
        with db_lock:
            cursor.execute("SELECT name, phone FROM users WHERE user_id=?", (user_id,))
            user_info = cursor.fetchone()
            
        if user_info:
            admin_msg = f"👤 YANGI NATIJA:\n{user_info[0]} ({user_info[1]})\nNatija: {score_percent}% ({status_text})\nInfo: {reason}"
            notify_admins(admin_msg)
            
    except Exception as e:
        print(f"Xatolik finish_test: {e}")

def notify_admins(text):
    with db_lock:
        cursor.execute("SELECT user_id FROM admins")
        admins = cursor.fetchall()
    for admin in admins:
        try:
            bot.send_message(admin[0], text)
        except:
            pass

# ================= BACKGROUND THREAD (VAQT NAZORATI) =================
def time_checker_loop():
    while True:
        try:
            current_time = int(time.time())
            limit = get_settings()[0]
            
            with db_lock:
                # Faqat active sessiyalarni olamiz (Index ishlaydi)
                cursor.execute("SELECT user_id, started_at, warning_sent FROM sessions WHERE active=1")
                active_tests = cursor.fetchall()
            
            for user in active_tests:
                u_id, start_t, warned = user
                elapsed = current_time - start_t
                remaining = limit - elapsed
                
                if remaining <= 0:
                    finish_test(u_id, forced=True, reason="Umumiy vaqt tugadi")
                elif remaining <= 60 and warned == 0:
                    try:
                        bot.send_message(u_id, "⚠️ DIQQAT! 1 daqiqa vaqt qoldi!", protect_content=True)
                        with db_lock:
                            cursor.execute("UPDATE sessions SET warning_sent=1 WHERE user_id=?", (u_id,))
                            conn.commit()
                    except:
                        pass
        except Exception as e:
            print(f"Timer error: {e}")
        time.sleep(10)

threading.Thread(target=time_checker_loop, daemon=True).start()

# ================= START va RO'YXATDAN O'TISH =================
@bot.message_handler(commands=['start'])
def start(msg):
    if check_flood(msg.chat.id): return
    
    # 2-BAND: /start BOSILSA — TEST DARHOL YOPILADI
    if check_active_session(msg.chat.id):
        finish_test(msg.chat.id, forced=True, reason="Test paytida /start bosildi (Qoida buzildi)")
        bot.send_message(msg.chat.id, "❌ Test davomida /start bosildi.\nTest bekor qilindi va natija hisoblandi.", protect_content=True)
        return

    # Agar oldin tugatgan bo'lsa
    with db_lock:
        cursor.execute("SELECT active FROM sessions WHERE user_id=?", (msg.chat.id,))
        row = cursor.fetchone()
    
    # Agar sessions jadvalida bor va active=0 bo'lsa -> Test topshirib bo'lgan
    if row is not None and row[0] == 0:
        bot.send_message(msg.chat.id, "Siz testni topshirib bo'lgansiz. Qayta topshirish mumkin emas.", protect_content=True)
        return

    bot.send_message(msg.chat.id, "👋 Assalomu alaykum! HR Botga xush kelibsiz.\nIltimos, ismingizni kiriting:", protect_content=True)
    bot.register_next_step_handler(msg, save_name)

def save_name(msg):
    if check_flood(msg.chat.id): return
    with db_lock:
        cursor.execute(
            "INSERT OR REPLACE INTO users(user_id, name, registered_at) VALUES (?,?,?)",
            (msg.chat.id, msg.text, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
    
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    btn = types.KeyboardButton("📞 Telefon raqamni yuborish", request_contact=True)
    markup.add(btn)
    
    bot.send_message(msg.chat.id, "📞 Telefon raqamingizni yuboring:", reply_markup=markup, protect_content=True)
    bot.register_next_step_handler(msg, save_phone)

def save_phone(msg):
    if check_flood(msg.chat.id): return
    phone = msg.contact.phone_number if msg.contact else msg.text
    with db_lock:
        cursor.execute(
            "UPDATE users SET phone=? WHERE user_id=?",
            (phone, msg.chat.id)
        )
        conn.commit()
    
    # 1-BAND: SESSION YARATISH
    create_session(msg.chat.id)
    
    limit = get_settings()[0]
    bot.send_message(
        msg.chat.id, 
        f"✅ Ro'yxatdan o'tdingiz.\n⏱ Test vaqti: {limit//60} daqiqa.\n⏱ Har bir savolga: {QUESTION_TIME} soniya.\nTest boshlanmoqda...", 
        reply_markup=types.ReplyKeyboardRemove(),
        protect_content=True
    )
    send_question(msg.chat.id)

# ================= TEST LOGIKASI (PRO ANTI-CHEAT) =================
def send_question(user_id):
    if not check_active_session(user_id): 
        return

    with db_lock:
        cursor.execute("SELECT started_at FROM sessions WHERE user_id=?", (user_id,))
        start_t = cursor.fetchone()[0]
    
    limit = get_settings()[0]
    remaining_seconds = limit - (int(time.time()) - start_t)
    
    if remaining_seconds <= 0:
        finish_test(user_id, forced=True, reason="Vaqt tugadi")
        return

    with db_lock:
        # Optimallashgan Random so'rov
        # Katta bazada (100k) RANDOM() sekin ishlashi mumkin, lekin subquery bilan bu yechim eng optimali
        cursor.execute("""
            SELECT * FROM questions 
            WHERE id NOT IN (SELECT question_id FROM answers WHERE user_id=?)
            ORDER BY RANDOM() LIMIT 1
        """, (user_id,))
        q = cursor.fetchone()

    if not q:
        finish_test(user_id, reason="Savollar tugadi") 
        return

    q_id = q[0]
    q_text = q[1]
    media_type = q[2]
    file_id = q[3]
    options = [q[4], q[5], q[6], q[7]]
    
    # 3-BAND: SAVOL VAQTINI BELGILASH
    user_question_start_time[user_id] = time.time()
    
    time_str = f"⏳ Qolgan vaqt: {remaining_seconds//60}:{remaining_seconds%60:02d}"
    limit_str = f"\n⚡️ Bu savolga {QUESTION_TIME} soniya vaqtingiz bor!"
    full_text = f"{time_str}{limit_str}\n\n❓ {q_text}\n\nA) {options[0]}\nB) {options[1]}\nC) {options[2]}\nD) {options[3]}"

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("A", "B", "C", "D")

    # ANTI-CHEAT: protect_content=True
    if media_type == 'photo':
        bot.send_photo(user_id, file_id, caption=full_text, reply_markup=markup, protect_content=True)
    elif media_type == 'video':
        bot.send_video(user_id, file_id, caption=full_text, reply_markup=markup, protect_content=True)
    else:
        bot.send_message(user_id, full_text, reply_markup=markup, protect_content=True)
        
    bot.register_next_step_handler_by_chat_id(user_id, save_answer, q_id, q[8])

def save_answer(msg, q_id, correct_option):
    user_id = msg.chat.id
    
    # 2-BAND: /start BOSILSA
    if msg.text == '/start':
        finish_test(user_id, forced=True, reason="Test paytida /start bosildi")
        bot.send_message(user_id, "❌ Test bekor qilindi.", protect_content=True)
        return

    if check_flood(user_id): 
        send_question(user_id) 
        return

    # 4-BAND: JAVOBNI QAYTA BOSISH BLOKI (DB check)
    with db_lock:
        cursor.execute("SELECT 1 FROM answers WHERE user_id=? AND question_id=?", (user_id, q_id))
        if cursor.fetchone():
            return # Allaqachon javob bergan

    # 3-BAND: VAQTNI TEKSHIRISH
    elapsed = time.time() - user_question_start_time.get(user_id, 0)
    if elapsed > QUESTION_TIME:
        bot.send_message(user_id, f"⏰ Vaqt tugadi! ({int(elapsed)}s > {QUESTION_TIME}s). Javob qabul qilinmadi.", protect_content=True)
        # Noto'g'ri javob sifatida belgilash yoki shunchaki o'tkazib yuborish
        # Biz bu yerda "noto'g'ri" deb belgilaymiz (bo'sh javob bilan)
        with db_lock:
            cursor.execute(
                "INSERT INTO answers(user_id, question_id, answer, is_correct, timestamp) VALUES (?,?,?,?,?)",
                (user_id, q_id, "TIMEOUT", 0, int(time.time()))
            )
            conn.commit()
        send_question(user_id)
        return

    user_ans = msg.text.upper() if msg.text else ""
    
    if user_ans not in ["A", "B", "C", "D"]:
        bot.send_message(user_id, "⚠️ Iltimos, variantlardan birini tanlang (A, B, C, D)", protect_content=True)
        bot.register_next_step_handler(msg, save_answer, q_id, correct_option)
        return

    is_correct = 1 if user_ans == correct_option else 0
    
    with db_lock:
        cursor.execute(
            "INSERT INTO answers(user_id, question_id, answer, is_correct, timestamp) VALUES (?,?,?,?,?)",
            (user_id, q_id, user_ans, is_correct, int(time.time()))
        )
        conn.commit()
    send_question(user_id)

# ================= ADMIN PANEL =================
@bot.message_handler(commands=['admin'])
def admin_panel(msg):
    if not is_admin(msg.chat.id):
        bot.send_message(msg.chat.id, "⛔ Siz admin emassiz")
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add("➕ Savol qo‘shish", "✏️ Savol tahrirlash")
    markup.add("🗑 Savol o‘chirish", "📋 Savollar ro‘yxati")
    markup.add("📊 Statistika", "📥 Excel Hisobot")
    markup.add("⚙️ Sozlamalar", "👥 Adminlar")
    markup.add("🗂 Baza ma’lumotlari", "📜 Loglar")
    markup.add("📈 Dashboard", "♻️ User testini reset qilish")
    markup.add("⛔ Testni majburiy to‘xtatish")
    markup.add("📉 Eng past natijalar", "🔥 Eng qiyin savollar")
    markup.add("🧹 Bazani tozalash", "📢 Xabar yuborish")
    markup.add("❌ Chiqish")

    bot.send_message(msg.chat.id, "🛠 KENGAYTIRILGAN ADMIN PANEL", reply_markup=markup)

# --- LOGLARNI KO'RISH ---
@bot.message_handler(func=lambda m: m.text == "📜 Loglar")
def view_logs(msg):
    if not is_admin(msg.chat.id): return
    
    # Audit loglar va Admin loglarni ko'rsatish
    with db_lock:
        cursor.execute("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 15")
        logs = cursor.fetchall()
    
    if not logs:
        bot.send_message(msg.chat.id, "📭 Loglar topilmadi")
        return
        
    text = "📜 SO'NGGI AUDIT LOGLAR:\n\n"
    for log in logs:
        text += f"🕒 {log[3]}\n👤 ID: {log[1]}\n📝 {log[2]}\n{'-'*15}\n"
    
    if len(text) > 4096:
        with open("logs.txt", "w", encoding='utf-8') as f:
            f.write(text)
        with open("logs.txt", "rb") as f:
            bot.send_document(msg.chat.id, f)
        os.remove("logs.txt")
    else:
        bot.send_message(msg.chat.id, text)

# --- 1. SOZLAMALAR ---
@bot.message_handler(func=lambda m: m.text == "⚙️ Sozlamalar")
def settings_menu(msg):
    if not is_admin(msg.chat.id): return
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("⏱ Vaqt limiti", "🎯 O'tish bali")
    markup.add("⬅️ Orqaga")
    bot.send_message(msg.chat.id, "Sozlamani tanlang:", reply_markup=markup)

@bot.message_handler(func=lambda m: m.text == "⏱ Vaqt limiti")
def set_time_limit(msg):
    bot.send_message(msg.chat.id, "Test vaqtini daqiqada kiriting:")
    bot.register_next_step_handler(msg, save_time_limit)

def save_time_limit(msg):
    try:
        minutes = int(msg.text)
        with db_lock:
            cursor.execute("UPDATE settings SET test_time=? WHERE id=1", (minutes*60,))
            conn.commit()
        log_admin_action(msg.chat.id, f"Vaqt limiti o'zgartirildi: {minutes} daqiqa")
        bot.send_message(msg.chat.id, "✅ Vaqt yangilandi")
    except:
        bot.send_message(msg.chat.id, "❌ Raqam kiriting")

@bot.message_handler(func=lambda m: m.text == "🎯 O'tish bali")
def set_pass_score(msg):
    bot.send_message(msg.chat.id, "O'tish balini foizda kiriting (masalan 70):")
    bot.register_next_step_handler(msg, save_pass_score)

def save_pass_score(msg):
    try:
        score = int(msg.text)
        with db_lock:
            cursor.execute("UPDATE settings SET pass_score=? WHERE id=1", (score,))
            conn.commit()
        log_admin_action(msg.chat.id, f"O'tish bali o'zgartirildi: {score}%")
        bot.send_message(msg.chat.id, "✅ O'tish bali yangilandi")
    except:
        bot.send_message(msg.chat.id, "❌ Raqam kiriting")

# --- 2. EXCEL HISOBOT ---
@bot.message_handler(func=lambda m: m.text == "📥 Excel Hisobot")
def export_excel_advanced(msg):
    if not is_admin(msg.chat.id): return
    
    bot.send_message(msg.chat.id, "⏳ Hisobot tayyorlanmoqda...")
    log_admin_action(msg.chat.id, "Excel hisobot yuklab olindi")
    
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Umumiy Natijalar"
    headers = ["ID", "Ism", "Telefon", "Ro'yxatdan o'tgan", "To'g'ri", "Jami Savol", "Foiz", "Status", "Start", "Active?"]
    ws1.append(headers)
    
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    with db_lock:
        cursor.execute("SELECT * FROM users")
        users = cursor.fetchall()
    
    settings = get_settings()
    pass_score = settings[1]
    
    for u in users:
        u_id, name, phone, reg_at = u
        with db_lock:
            cursor.execute("SELECT COUNT(*) FROM answers WHERE user_id=? AND is_correct=1", (u_id,))
            correct = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM answers WHERE user_id=?", (u_id,))
            total_ans = cursor.fetchone()[0]
        
        percent = 0
        if total_ans > 0:
            percent = int((correct / total_ans) * 100)
            
        status = "O'TDI" if percent >= pass_score else "YIQILDI"
        with db_lock:
            cursor.execute("SELECT started_at, active FROM sessions WHERE user_id=?", (u_id,))
            ts = cursor.fetchone()
        
        start_str = datetime.fromtimestamp(ts[0]).strftime('%H:%M:%S') if ts else "-"
        active_str = "Ha" if ts and ts[1] else "Yo'q"
        ws1.append([u_id, name, phone, reg_at, correct, total_ans, f"{percent}%", status, start_str, active_str])

    ws2 = wb.create_sheet("Batafsil Javoblar")
    ws2.append(["User ID", "Ism", "Savol ID", "Berilgan Javob", "To'g'ri Javob", "Natija"])
    with db_lock:
        cursor.execute("""
            SELECT u.user_id, u.name, q.id, a.answer, q.correct, a.is_correct
            FROM answers a
            JOIN users u ON u.user_id = a.user_id
            JOIN questions q ON q.id = a.question_id
        """)
        details = cursor.fetchall()
    for row in details:
        res_str = "✅" if row[5] == 1 else "❌"
        ws2.append([row[0], row[1], row[2], row[3], row[4], res_str])

    filename = f"Hisobot_{int(time.time())}.xlsx"
    wb.save(filename)
    with open(filename, "rb") as f:
        bot.send_document(msg.chat.id, f, caption="📊 To'liq hisobot")
    os.remove(filename)

# --- 3. SAVOL QO'SHISH ---
@bot.message_handler(func=lambda m: m.text == "➕ Savol qo‘shish")
def add_question_start(msg):
    if not is_admin(msg.chat.id): return
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add("Matn", "Rasm", "Video")
    bot.send_message(msg.chat.id, "Savol turini tanlang:", reply_markup=markup)
    bot.register_next_step_handler(msg, ask_question_content)

def ask_question_content(msg):
    q_type = "text"
    if msg.text == "Rasm": q_type = "photo"
    elif msg.text == "Video": q_type = "video"
    bot.send_message(msg.chat.id, f"Savol matnini kiriting (yoki {q_type} bo'lsa caption uchun matn):")
    bot.register_next_step_handler(msg, save_question_text_media, q_type)

def save_question_text_media(msg, q_type):
    q_text = msg.text
    if q_type == "text":
        ask_variants(msg, q_text, "text", None)
    else:
        bot.send_message(msg.chat.id, f"Endi {q_type}ni yuboring:")
        bot.register_next_step_handler(msg, save_media_file, q_text, q_type)

def save_media_file(msg, q_text, q_type):
    file_id = None
    if q_type == "photo" and msg.photo:
        file_id = msg.photo[-1].file_id
    elif q_type == "video" and msg.video:
        file_id = msg.video.file_id
    else:
        bot.send_message(msg.chat.id, "❌ Noto'g'ri format.")
        return
    ask_variants(msg, q_text, q_type, file_id)

def ask_variants(msg, q_text, q_type, file_id):
    bot.send_message(msg.chat.id, "Variant A:")
    bot.register_next_step_handler(msg, save_a, q_text, q_type, file_id)

def save_a(msg, q_text, q_type, file_id):
    a = msg.text
    bot.send_message(msg.chat.id, "Variant B:")
    bot.register_next_step_handler(msg, save_b, q_text, q_type, file_id, a)

def save_b(msg, q_text, q_type, file_id, a):
    b = msg.text
    bot.send_message(msg.chat.id, "Variant C:")
    bot.register_next_step_handler(msg, save_c, q_text, q_type, file_id, a, b)

def save_c(msg, q_text, q_type, file_id, a, b):
    c = msg.text
    bot.send_message(msg.chat.id, "Variant D:")
    bot.register_next_step_handler(msg, save_d, q_text, q_type, file_id, a, b, c)

def save_d(msg, q_text, q_type, file_id, a, b, c):
    d = msg.text
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add("A", "B", "C", "D")
    bot.send_message(msg.chat.id, "To'g'ri javob:", reply_markup=markup)
    bot.register_next_step_handler(msg, commit_question, q_text, q_type, file_id, a, b, c, d)

def commit_question(msg, q_text, q_type, file_id, a, b, c, d):
    correct = msg.text.upper()
    if correct not in ["A", "B", "C", "D"]:
        bot.send_message(msg.chat.id, "Xato! A, B, C yoki D tanlang.")
        return
    
    with db_lock:
        cursor.execute("""
            INSERT INTO questions (question, media_type, file_id, a, b, c, d, correct)
            VALUES (?,?,?,?,?,?,?,?)
        """, (q_text, q_type, file_id, a, b, c, d, correct))
        conn.commit()
    log_admin_action(msg.chat.id, f"Yangi savol qo'shildi: {q_text[:30]}...")
    bot.send_message(msg.chat.id, "✅ Savol saqlandi!", reply_markup=types.ReplyKeyboardRemove())
    admin_panel(msg)

# --- ADMINLAR ---
@bot.message_handler(func=lambda m: m.text == "👥 Adminlar")
def manage_admins(msg):
    if not is_admin(msg.chat.id): return
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("➕ Admin qo'shish", "🗑 Admin o'chirish")
    markup.add("📜 Adminlar ro'yxati", "⬅️ Orqaga")
    bot.send_message(msg.chat.id, "Adminlar bo'limi:", reply_markup=markup)

@bot.message_handler(func=lambda m: m.text == "➕ Admin qo'shish")
def add_admin_ask(msg):
    bot.send_message(msg.chat.id, "Yangi admin ID:")
    bot.register_next_step_handler(msg, save_new_admin)

def save_new_admin(msg):
    try:
        new_id = int(msg.text)
        with db_lock:
            cursor.execute("INSERT OR IGNORE INTO admins (user_id, added_by) VALUES (?,?)", (new_id, msg.chat.id))
            conn.commit()
        log_admin_action(msg.chat.id, f"Yangi admin qo'shildi: {new_id}")
        bot.send_message(msg.chat.id, "✅ Admin qo'shildi.")
    except:
        bot.send_message(msg.chat.id, "❌ ID raqam bo'lishi kerak.")

@bot.message_handler(func=lambda m: m.text == "🗑 Admin o'chirish")
def del_admin_ask(msg):
    bot.send_message(msg.chat.id, "O'chirish uchun ID:")
    bot.register_next_step_handler(msg, delete_admin_do)

def delete_admin_do(msg):
    try:
        del_id = int(msg.text)
        if del_id == MAIN_ADMIN_ID:
            bot.send_message(msg.chat.id, "❌ Asosiy adminni o'chira olmaysiz!")
            return
        with db_lock:
            cursor.execute("DELETE FROM admins WHERE user_id=?", (del_id,))
            conn.commit()
        log_admin_action(msg.chat.id, f"Admin o'chirildi: {del_id}")
        bot.send_message(msg.chat.id, "✅ Admin o'chirildi.")
    except:
        bot.send_message(msg.chat.id, "Xato ID")

@bot.message_handler(func=lambda m: m.text == "📜 Adminlar ro'yxati")
def list_admins(msg):
    with db_lock:
        cursor.execute("SELECT user_id FROM admins")
        admins = cursor.fetchall()
    text = "👥 Adminlar:\n"
    for a in admins:
        text += f"- {a[0]}\n"
    bot.send_message(msg.chat.id, text)

# --- BAZA ---
@bot.message_handler(func=lambda m: m.text == "🗂 Baza ma’lumotlari")
def db_menu(msg):
    if not is_admin(msg.chat.id): return
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("👤 Users", "❓ Questions")
    markup.add("📝 Answers", "⏱ Sessions")
    markup.add("👥 Adminlar", "⬅️ Orqaga")
    bot.send_message(msg.chat.id, "Qaysi jadval?", reply_markup=markup)

@bot.message_handler(func=lambda m: m.text == "👤 Users")
def show_users_table(msg):
    with db_lock:
        cursor.execute("SELECT * FROM users")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(msg.chat.id, "Bo'sh")
        return
    text = "👤 USERS:\n"
    for r in rows:
        text += f"ID: {r[0]} | {r[1]} | {r[2]}\n"
        if len(text) > 3500:
            bot.send_message(msg.chat.id, text)
            text = ""
    if text: bot.send_message(msg.chat.id, text)

@bot.message_handler(func=lambda m: m.text == "❓ Questions")
def show_questions_table(msg):
    with db_lock:
        cursor.execute("SELECT id, question, correct FROM questions")
        rows = cursor.fetchall()
    text = "❓ SAVOLLAR:\n"
    for r in rows:
        text += f"{r[0]}. {r[1]} ({r[2]})\n"
        if len(text) > 3500:
            bot.send_message(msg.chat.id, text)
            text = ""
    if text: bot.send_message(msg.chat.id, text)

@bot.message_handler(func=lambda m: m.text == "⏱ Sessions")
def show_sessions(msg):
    with db_lock:
        cursor.execute("SELECT * FROM sessions")
        rows = cursor.fetchall()
    text = "⏱ SESSIONS:\n"
    for r in rows:
        text += f"UID: {r[0]} | Act: {r[3]}\n"
    bot.send_message(msg.chat.id, text[:4096])

@bot.message_handler(func=lambda m: m.text == "📈 Dashboard")
def dashboard(msg):
    if not is_admin(msg.chat.id): return
    with db_lock:
        cursor.execute("SELECT COUNT(*) FROM sessions WHERE active=1")
        active = cursor.fetchone()[0]
        cursor.execute("SELECT AVG(is_correct) FROM answers")
        avg = cursor.fetchone()[0]
    avg_percent = int(avg*100) if avg else 0
    text = f"📊 DASHBOARD\n🟢 Aktiv: {active}\n📈 O'rtacha: {avg_percent}%"
    bot.send_message(msg.chat.id, text)

@bot.message_handler(func=lambda m: m.text == "♻️ User testini reset qilish")
def reset_user_test(msg):
    if not is_admin(msg.chat.id): return
    bot.send_message(msg.chat.id, "User ID:")
    bot.register_next_step_handler(msg, do_reset_user)

def do_reset_user(msg):
    try:
        uid = int(msg.text)
        with db_lock:
            cursor.execute("DELETE FROM answers WHERE user_id=?", (uid,))
            cursor.execute("DELETE FROM sessions WHERE user_id=?", (uid,))
            conn.commit()
        log_admin_action(msg.chat.id, f"User reset qilindi: {uid}")
        bot.send_message(msg.chat.id, f"✅ {uid} reset qilindi")
    except:
        bot.send_message(msg.chat.id, "Xato ID")

@bot.message_handler(func=lambda m: m.text == "⛔ Testni majburiy to‘xtatish")
def force_finish_admin(msg):
    bot.send_message(msg.chat.id, "User ID:")
    bot.register_next_step_handler(msg, force_finish_do)

def force_finish_do(msg):
    try:
        uid = int(msg.text)
        finish_test(uid, forced=True, reason="Admin tomonidan majburiy to'xtatildi")
        log_admin_action(msg.chat.id, f"User majburiy to'xtatildi: {uid}")
        bot.send_message(msg.chat.id, "⛔ Test to'xtatildi")
    except:
        bot.send_message(msg.chat.id, "Xato ID")

@bot.message_handler(func=lambda m: m.text == "📉 Eng past natijalar")
def worst_users(msg):
    with db_lock:
        cursor.execute("SELECT u.name, AVG(a.is_correct)*100 as score FROM answers a JOIN users u ON u.user_id=a.user_id GROUP BY a.user_id ORDER BY score ASC LIMIT 5")
        rows = cursor.fetchall()
    text = "📉 ENG PAST:\n"
    for r in rows: text += f"{r[0]} — {int(r[1])}%\n"
    bot.send_message(msg.chat.id, text)

@bot.message_handler(func=lambda m: m.text == "🔥 Eng qiyin savollar")
def hardest_questions(msg):
    with db_lock:
        cursor.execute("SELECT q.id, q.question, COUNT(*) c FROM answers a JOIN questions q ON q.id=a.question_id WHERE a.is_correct=0 GROUP BY q.id ORDER BY c DESC LIMIT 5")
        rows = cursor.fetchall()
    text = "🔥 ENG QIYIN:\n"
    for r in rows: text += f"ID {r[0]} ({r[2]} xato): {r[1]}\n"
    bot.send_message(msg.chat.id, text[:4096])

@bot.message_handler(func=lambda m: m.text == "📢 Xabar yuborish")
def broadcast(msg):
    if not is_admin(msg.chat.id): return
    bot.send_message(msg.chat.id, "Matn:")
    bot.register_next_step_handler(msg, send_broadcast)

def send_broadcast(msg):
    with db_lock:
        cursor.execute("SELECT user_id FROM users")
        users = cursor.fetchall()
    count = 0
    for u in users:
        try:
            bot.send_message(u[0], msg.text)
            count += 1
        except: pass
    log_admin_action(msg.chat.id, f"Xabar yuborildi: {count} kishiga")
    bot.send_message(msg.chat.id, f"📢 {count} kishiga bordi")

@bot.message_handler(func=lambda m: m.text == "⬅️ Orqaga")
def back_to_main(msg):
    admin_panel(msg)

@bot.message_handler(func=lambda m: m.text == "📋 Savollar ro‘yxati")
def list_questions(msg):
    if not is_admin(msg.chat.id): return
    with db_lock:
        cursor.execute("SELECT id, question, media_type, correct FROM questions")
        rows = cursor.fetchall()
    text = "📋 Savollar:\n"
    for r in rows:
        text += f"{r[0]}. {r[1]} ({r[3]})\n"
    bot.send_message(msg.chat.id, text[:4096])

@bot.message_handler(func=lambda m: m.text == "📊 Statistika")
def stats(msg):
    if not is_admin(msg.chat.id): return
    with db_lock:
        cursor.execute("SELECT COUNT(*) FROM users")
        u_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM sessions WHERE active=0")
        finished_count = cursor.fetchone()[0]
    settings = get_settings()
    bot.send_message(msg.chat.id, f"📊 Statistika:\n👥 Userlar: {u_count}\n🏁 Tugatganlar: {finished_count}")

@bot.message_handler(func=lambda m: m.text == "🗑 Savol o‘chirish")
def delete_question_start(msg):
    if not is_admin(msg.chat.id): return
    bot.send_message(msg.chat.id, "ID:")
    def delete_op(m):
        with db_lock:
            cursor.execute("DELETE FROM questions WHERE id=?", (m.text,))
            conn.commit()
        log_admin_action(msg.chat.id, f"Savol o'chirildi: {m.text}")
        bot.send_message(m.chat.id, "O'chirildi")
    bot.register_next_step_handler(msg, delete_op)

@bot.message_handler(func=lambda m: m.text == "🧹 Bazani tozalash")
def clear_db(msg):
    if not is_admin(msg.chat.id): return
    with db_lock:
        cursor.execute("DELETE FROM users")
        cursor.execute("DELETE FROM answers")
        cursor.execute("DELETE FROM sessions")
        conn.commit()
    log_admin_action(msg.chat.id, "Baza tozalandi")
    bot.send_message(msg.chat.id, "Baza tozalandi.")

@bot.message_handler(func=lambda m: m.text == "❌ Chiqish")
def exit_admin(msg):
    bot.send_message(msg.chat.id, "Yopildi.", reply_markup=types.ReplyKeyboardRemove())

# ================= RUN =================
print("🤖 HR Bot V2 (Pro Security) ishga tushdi...")
try:
    bot.polling(non_stop=True)
except Exception as e:
    print(f"Error: {e}")